import csv
import io
import json
from uuid import uuid4

from fastapi import APIRouter, Depends, File, HTTPException, UploadFile, BackgroundTasks
from sqlalchemy.ext.asyncio import AsyncSession
import openpyxl

from app.core.dependencies import get_db
from app.models.ingestion import RawPayload, JobRun
from ingestion.plugins.registry import SOURCES
from ingestion.core.session import SessionManager
from ingestion.core.normalization import NormalizationEngine
from ingestion.core.dedupe import DeduplicationEngine
from ingestion.core.geocoder import Geocoder

router = APIRouter()

async def process_payload(source: str, payload_id: str, db: AsyncSession):
    # This runs as a background task to process the payload
    payload = await db.get(RawPayload, payload_id)
    if not payload:
        return

    PluginClass = SOURCES.get(source)
    if not PluginClass:
        return

    session_manager = SessionManager(db)
    plugin = PluginClass(db=db, session_manager=session_manager)
    
    # 1. Normalize
    listings = plugin.normalize(payload)
    
    # 2. Geocode & Dedupe
    geocoder = Geocoder(db)
    dedupe = DeduplicationEngine(db)
    from app.models.property import Property
    
    saved = 0
    for listing in listings:
        # Geocode
        geo = await geocoder.geocode(listing)
        if geo:
            listing.latitude = geo.latitude
            listing.longitude = geo.longitude
            
        fingerprint = dedupe.generate_fingerprint(listing)
        prop_id = await dedupe.check_duplicate(listing, fingerprint)
        
        if not prop_id:
            # Create new property
            prop = Property(
                source=listing.source,
                external_id=listing.external_id,
                fingerprint=fingerprint,
                title=listing.title,
                property_type=listing.property_type,
                price=listing.price,
                area_sqft=listing.area_sqft,
                bedrooms=listing.bedrooms,
                bathrooms=listing.bathrooms,
                address=listing.address,
                locality=listing.locality,
                city=listing.city,
                latitude=listing.latitude,
                longitude=listing.longitude,
                listing_url=listing.listing_url,
                confidence_score=80 if source in ["builder_direct"] else 50
            )
            if listing.latitude and listing.longitude:
                prop.location = f"POINT({listing.longitude} {listing.latitude})"
                
            db.add(prop)
            await db.commit()
            saved += 1
            
    # Note: In production we would update JobRun here
    print(f"Processed {len(listings)} listings, saved {saved} new properties.")


@router.post("/csv")
async def ingest_csv(
    background_tasks: BackgroundTasks,
    file: UploadFile = File(...),
    db: AsyncSession = Depends(get_db)
):
    if not file.filename.endswith('.csv'):
        raise HTTPException(status_code=400, detail="Must be a CSV file")
        
    content = await file.read()
    reader = csv.DictReader(io.StringIO(content.decode("utf-8")))
    rows = [row for row in reader]
    
    payload = RawPayload(
        source="csv_upload",
        endpoint=f"file_upload:{file.filename}",
        payload_json={"rows": rows}
    )
    db.add(payload)
    await db.commit()
    
    background_tasks.add_task(process_payload, "csv_upload", payload.id, db)
    return {"status": "accepted", "message": f"Processing {len(rows)} rows from {file.filename}"}


@router.post("/excel")
async def ingest_excel(
    background_tasks: BackgroundTasks,
    file: UploadFile = File(...),
    db: AsyncSession = Depends(get_db)
):
    if not file.filename.endswith(('.xls', '.xlsx')):
        raise HTTPException(status_code=400, detail="Must be an Excel file")
        
    content = await file.read()
    wb = openpyxl.load_workbook(filename=io.BytesIO(content), data_only=True)
    ws = wb.active
    
    rows = []
    headers = [cell.value for cell in ws[1]]
    for row in ws.iter_rows(min_row=2, values_only=True):
        if any(row):
            rows.append(dict(zip(headers, row)))
            
    payload = RawPayload(
        source="excel_upload",
        endpoint=f"file_upload:{file.filename}",
        payload_json={"rows": rows}
    )
    db.add(payload)
    await db.commit()
    
    background_tasks.add_task(process_payload, "excel_upload", payload.id, db)
    return {"status": "accepted", "message": f"Processing {len(rows)} rows from {file.filename}"}


@router.post("/json")
async def ingest_json(
    background_tasks: BackgroundTasks,
    file: UploadFile = File(...),
    db: AsyncSession = Depends(get_db)
):
    if not file.filename.endswith('.json'):
        raise HTTPException(status_code=400, detail="Must be a JSON file")
        
    content = await file.read()
    data = json.loads(content)
    
    payload = RawPayload(
        source="json_upload",
        endpoint=f"file_upload:{file.filename}",
        payload_json=data
    )
    db.add(payload)
    await db.commit()
    
    background_tasks.add_task(process_payload, "json_upload", payload.id, db)
    return {"status": "accepted", "message": f"Processing JSON from {file.filename}"}


@router.post("/upload")
async def inventory_upload(
    payload_data: dict,
    background_tasks: BackgroundTasks,
    db: AsyncSession = Depends(get_db)
):
    payload = RawPayload(
        source="manual_upload",
        endpoint="api:manual_upload",
        payload_json=payload_data
    )
    db.add(payload)
    await db.commit()
    
    background_tasks.add_task(process_payload, "manual_upload", payload.id, db)
    return {"status": "accepted", "message": "Manual upload processing"}
